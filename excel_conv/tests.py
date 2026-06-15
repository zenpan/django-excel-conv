import shutil
import tempfile
from io import BytesIO
from pathlib import Path
from unittest import mock

from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from django_excel import __version__

from excel_conv.lib.convert import convert_sheet, _format_judgment, _parse_filing_fields
from excel_conv.lib import nysc
from excel_conv.lib.nysc import extract_nysc_records, _parse_nysc_address
from excel_conv.lib.sources import convert_job, detect_source, source_label
from excel_conv.models import ConvJob


TEST_MEDIA_ROOT = Path(tempfile.mkdtemp(prefix="excel-conv-test-media-"))


def make_source_workbook():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "No."
    worksheet["B3"] = "DOE, JOHN Q"
    worksheet["C3"] = "123 Main St\nNewark, NJ 07102\nEssex County"
    worksheet["D3"] = (
        "Filing Date:1/1/2026\nAmount:$10,329\nCIVIL JUDGMENT\n"
        "Filing Number:ABC123\nFiling Date:1/2/2026\nBook/Page:100/200\n"
        "Filing Office:Test Court"
    )
    worksheet["E3"] = "Creditor LLC"
    worksheet["A6"] = "Permissible Use:"

    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    stream.seek(0)
    return stream.getvalue()


def make_multi_sheet_source_workbook():
    """A workbook whose data is NOT on the active sheet, mirroring the Florida
    exports: an empty 'Sheet1' is active and the records live on the
    'Public Records Results List' sheet."""
    workbook = Workbook()
    workbook.active.title = "Sheet1"  # empty stub, active by default
    data = workbook.create_sheet("Public Records Results List")
    data["A1"] = "No."
    data["B3"] = "DOE, JANE Q"
    data["C3"] = "123 Main St\nMiami, FL 33101-1234\nMiami-Dade County"
    data["D3"] = "Filing Date:1/12/2026\nAmount:$2,251\nSMALL CLAIMS JUDGMENT"
    data["E3"] = "Creditor LLC"
    data["A6"] = "Permissible Use:"

    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    stream.seek(0)
    return stream.getvalue()


def make_workbook_with_bad_row():
    """One valid record plus a record that has a creditor but an EMPTY address
    cell (the job-516 case). Pre-fix this raised AttributeError and aborted the
    whole conversion; it must now be skipped, leaving the valid row converted."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "No."
    # valid record (row 3)
    worksheet["B3"] = "DOE, JOHN Q"
    worksheet["C3"] = "123 Main St\nNewark, NJ 07102\nEssex County"
    worksheet["D3"] = "Amount:$500"
    worksheet["E3"] = "Creditor LLC"
    # malformed record (row 4): creditor present, address missing
    worksheet["B4"] = "DESMOND, THOMAS"
    worksheet["E4"] = "LVNV FUNDING LLC"
    worksheet["A7"] = "Permissible Use:"  # ending_row = 7 - 3 = 4

    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    stream.seek(0)
    return stream.getvalue()


def make_workbook_with_codefendants():
    """A claim with three co-debtors in the LexisNexis layout: the first row
    carries the creditor + amount; the following co-defendant rows have only a
    name + address (blank No. / Filing / Creditor). The primary is a company."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "No."
    # primary debtor (a company) — carries creditor + amount
    worksheet["A3"] = "525."
    worksheet["B3"] = "EARL TRANSPORT SERVICE LLC"
    worksheet["C3"] = "18738 Deer Tracks Loop\nLutz, FL 33558-8487\nPasco County"
    worksheet["D3"] = "Filing Date:1/12/2026\nAmount:$21,380\nCIVIL JUDGMENT"
    worksheet["E3"] = "AMERICAN EXPRESS NATIONAL BANK"
    # co-defendant 1 (continuation row: name + address only)
    worksheet["B4"] = "ROBLES, EARL\nLexID(sm):\n189601170822"
    worksheet["C4"] = "18738 Deer Tracks Loop\nLutz, FL 33558-8487\nPasco County"
    # co-defendant 2
    worksheet["B5"] = "SOTO, EARL ROBLES\nLexID(sm):\n189601170822"
    worksheet["C5"] = "18738 Deer Tracks Loop\nLutz, FL 33558-8487\nPasco County"
    worksheet["A8"] = "Permissible Use:"  # ending_row = 8 - 3 = 5 -> rows 3,4,5

    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    stream.seek(0)
    return stream.getvalue()


@override_settings(MEDIA_ROOT=TEST_MEDIA_ROOT)
class ConversionTests(TestCase):
    @classmethod
    def tearDownClass(cls):
        super().tearDownClass()
        shutil.rmtree(TEST_MEDIA_ROOT, ignore_errors=True)

    def setUp(self):
        (TEST_MEDIA_ROOT / "excel_files").mkdir(parents=True, exist_ok=True)
        (TEST_MEDIA_ROOT / "conv_files").mkdir(parents=True, exist_ok=True)

    def test_convert_sheet_creates_mail_merge_workbook(self):
        job = ConvJob.objects.create(
            excel_file=SimpleUploadedFile(
                "source.xlsx",
                make_source_workbook(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        )

        self.assertTrue(convert_sheet(job))

        job.refresh_from_db()
        self.assertTrue(job.success)
        self.assertIsNotNone(job.conv_at)

        converted = load_workbook(TEST_MEDIA_ROOT / job.conv_file.name)
        worksheet = converted.active
        self.assertEqual(
            [worksheet[f"{column}1"].value for column in "ABCDEFGHIJKLM"],
            [
                "name", "ADDRESS_1", "City", "State", "Zip", "Creditor", "Judgment",
                "FilingDate", "JudgmentType", "FilingNumber", "FilingDate2",
                "BookPage", "FilingOffice",
            ],
        )
        self.assertEqual(
            [worksheet[f"{column}2"].value for column in "ABCDEFGHIJKLM"],
            [
                "JOHN Q DOE", "123 Main St", "Newark", "NJ", "07102", "Creditor LLC",
                "$10,329.00", "1/1/2026", "CIVIL JUDGMENT", "ABC123", "1/2/2026",
                "100/200", "Test Court",
            ],
        )
        converted.close()

    def test_convert_finds_data_sheet_when_not_active(self):
        # Regression for the Florida exports (the production /convert/517
        # incident): records live on the 'Public Records Results List' sheet
        # while an empty 'Sheet1' is active. The converter must find the data
        # sheet rather than blindly reading workbook.active.
        job = ConvJob.objects.create(
            excel_file=SimpleUploadedFile(
                "FL_SOUTH.xlsx", make_multi_sheet_source_workbook()
            )
        )

        self.assertTrue(convert_sheet(job))

        job.refresh_from_db()
        self.assertTrue(job.success)
        converted = load_workbook(TEST_MEDIA_ROOT / job.conv_file.name)
        worksheet = converted.active
        self.assertEqual(
            [worksheet[f"{column}2"].value for column in "ABCDEFG"],
            ["JANE Q DOE", "123 Main St", "Miami", "FL", "33101-1234", "Creditor LLC", "$2,251.00"],
        )
        converted.close()

    def test_format_judgment_variants(self):
        self.assertEqual(
            _format_judgment("Filing Date:1/1/2026\nAmount:$7,798\nCIVIL JUDGMENT"),
            "$7,798.00",
        )
        self.assertEqual(_format_judgment("Amount:$122,657"), "$122,657.00")
        self.assertEqual(_format_judgment("Amount:$1,234.56"), "$1,234.56")
        self.assertEqual(_format_judgment("Filing Date:1/1/2026\nno amount here"), "")
        self.assertEqual(_format_judgment(None), "")

    def test_parse_filing_fields(self):
        full = (
            "Filing Date:1/12/2026\nAmount:$7,798\nSMALL CLAIMS JUDGMENT\n"
            "Filing Number:2025149620SP21\nFiling Date:1/13/2026\n"
            "Book/Page:35116/2236\nFiling Office:CIRCUIT COURT - MIAMI, FL"
        )
        self.assertEqual(
            _parse_filing_fields(full),
            {
                "FilingDate": "1/12/2026",
                "JudgmentType": "SMALL CLAIMS JUDGMENT",
                "FilingNumber": "2025149620SP21",
                "FilingDate2": "1/13/2026",
                "BookPage": "35116/2236",
                "FilingOffice": "CIRCUIT COURT - MIAMI, FL",
            },
        )
        # Missing fields (no Book/Page) come back empty, not absent.
        no_bookpage = (
            "Filing Date:3/13/2023\nAmount:$10,329\nCIVIL JUDGMENT\n"
            "Filing Number:COWE22003271\nFiling Date:3/13/2023\n"
            "Filing Office:BROWARD CIRCUIT COURT, FL"
        )
        parsed = _parse_filing_fields(no_bookpage)
        self.assertEqual(parsed["BookPage"], "")
        self.assertEqual(parsed["FilingNumber"], "COWE22003271")
        # Empty / missing input -> every field blank.
        self.assertEqual(set(_parse_filing_fields(None).values()), {""})

    def test_convert_skips_row_with_missing_address(self):
        # Regression for job 516: a record with a creditor but an empty address
        # used to crash the whole conversion. It must now be skipped while the
        # valid records still convert.
        job = ConvJob.objects.create(
            excel_file=SimpleUploadedFile("bad_row.xlsx", make_workbook_with_bad_row())
        )

        self.assertTrue(convert_sheet(job))

        job.refresh_from_db()
        self.assertTrue(job.success)
        converted = load_workbook(TEST_MEDIA_ROOT / job.conv_file.name)
        worksheet = converted.active
        # header + exactly one valid row; the malformed row was skipped
        self.assertEqual(worksheet.max_row, 2)
        self.assertEqual(worksheet["A2"].value, "JOHN Q DOE")
        converted.close()

    def test_convert_captures_co_defendants(self):
        # Co-defendants/co-debtors (LexisNexis continuation rows with a blank
        # creditor) must each become their own row, inheriting the claim's
        # creditor + judgment; company names are kept as-is.
        job = ConvJob.objects.create(
            excel_file=SimpleUploadedFile("codef.xlsx", make_workbook_with_codefendants())
        )

        self.assertTrue(convert_sheet(job))

        job.refresh_from_db()
        self.assertTrue(job.success)
        converted = load_workbook(TEST_MEDIA_ROOT / job.conv_file.name)
        worksheet = converted.active
        # header + one row per co-debtor
        self.assertEqual(worksheet.max_row, 4)
        self.assertEqual(
            [worksheet.cell(row=r, column=1).value for r in (2, 3, 4)],
            ["EARL TRANSPORT SERVICE LLC", "EARL ROBLES", "EARL ROBLES SOTO"],
        )
        # every co-debtor shares the claim's creditor + judgment amount
        for r in (2, 3, 4):
            self.assertEqual(worksheet.cell(row=r, column=6).value, "AMERICAN EXPRESS NATIONAL BANK")
            self.assertEqual(worksheet.cell(row=r, column=7).value, "$21,380.00")
        converted.close()


@override_settings(MEDIA_ROOT=TEST_MEDIA_ROOT)
class ViewTests(TestCase):
    @classmethod
    def tearDownClass(cls):
        super().tearDownClass()
        shutil.rmtree(TEST_MEDIA_ROOT, ignore_errors=True)

    def setUp(self):
        (TEST_MEDIA_ROOT / "excel_files").mkdir(parents=True, exist_ok=True)
        (TEST_MEDIA_ROOT / "conv_files").mkdir(parents=True, exist_ok=True)
        self.user = User.objects.create_user(username="tester", password="password")

    def test_public_pages_render(self):
        for url_name in ("index", "about", "contact", "help"):
            with self.subTest(url_name=url_name):
                response = self.client.get(reverse(url_name))
                self.assertEqual(response.status_code, 200)

    def test_pages_show_copyright_and_version_footer(self):
        response = self.client.get(reverse("index"))
        self.assertContains(response, "ZenPan Technology Solutions")
        self.assertContains(response, __version__)

    def test_job_pages_require_login(self):
        for url_name, args in (
            ("jobs", None),
            ("upload", None),
            ("convert", [1]),
            ("delete", [1]),
        ):
            with self.subTest(url_name=url_name):
                response = self.client.get(reverse(url_name, args=args))
                self.assertEqual(response.status_code, 302)
                self.assertIn(reverse("login"), response["Location"])

    def test_authenticated_upload_creates_job(self):
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("upload"),
            {
                "excel_file": SimpleUploadedFile(
                    "source.xlsx",
                    make_source_workbook(),
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        self.assertRedirects(response, reverse("jobs"))
        self.assertEqual(ConvJob.objects.count(), 1)

    def test_jobs_page_renders_success_state(self):
        self.client.force_login(self.user)
        ConvJob.objects.create(
            excel_file=SimpleUploadedFile("source.xlsx", make_source_workbook()),
            conv_file="conv_files/source_converted.xlsx",
            success=True,
        )

        response = self.client.get(reverse("jobs"))

        self.assertContains(response, "Completed")
        self.assertContains(response, "Done")

    def test_convert_missing_job_returns_404_not_500(self):
        # Regression: /convert/<missing id> used to raise DoesNotExist -> HTTP 500.
        # It must be 404.
        self.client.force_login(self.user)
        response = self.client.get(reverse("convert", args=[999999]))
        self.assertEqual(response.status_code, 404)

    def test_delete_missing_job_returns_404_not_500(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse("delete", args=[517]))
        self.assertEqual(response.status_code, 404)

    def test_convert_existing_job_succeeds(self):
        self.client.force_login(self.user)
        job = ConvJob.objects.create(
            excel_file=SimpleUploadedFile("source.xlsx", make_source_workbook()),
        )

        response = self.client.get(reverse("convert", args=[job.id]))

        self.assertRedirects(response, reverse("jobs"))
        job.refresh_from_db()
        self.assertTrue(job.success)

    def test_convert_unparseable_file_fails_gracefully(self):
        # Regression for the production /convert/517 incident: an Excel file
        # without the expected "No." / "Permissible Use:" markers (e.g. the
        # Florida export) raised UnboundLocalError -> HTTP 500. It must now
        # redirect with an error message instead of crashing.
        self.client.force_login(self.user)
        workbook = Workbook()
        workbook.active["A1"] = "completely different layout"
        stream = BytesIO()
        workbook.save(stream)
        workbook.close()
        stream.seek(0)
        job = ConvJob.objects.create(
            excel_file=SimpleUploadedFile("FL_SOUTH.xlsx", stream.getvalue()),
        )

        response = self.client.get(reverse("convert", args=[job.id]))

        self.assertRedirects(response, reverse("jobs"))
        job.refresh_from_db()
        self.assertFalse(job.success)

    def test_download_requires_login(self):
        job = ConvJob.objects.create(
            excel_file=SimpleUploadedFile("source.xlsx", make_source_workbook()),
        )
        response = self.client.get(reverse("download", args=[job.id, "source"]))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse("login"), response["Location"])

    def test_download_converted_file_for_logged_in_user(self):
        self.client.force_login(self.user)
        job = ConvJob.objects.create(
            excel_file=SimpleUploadedFile("source.xlsx", make_source_workbook()),
        )
        self.assertTrue(convert_sheet(job))

        response = self.client.get(reverse("download", args=[job.id, "converted"]))

        self.assertEqual(response.status_code, 200)
        self.assertIn("attachment", response["Content-Disposition"])

    def test_download_missing_converted_file_returns_404(self):
        self.client.force_login(self.user)
        job = ConvJob.objects.create(  # never converted -> no conv_file
            excel_file=SimpleUploadedFile("source.xlsx", make_source_workbook()),
        )
        response = self.client.get(reverse("download", args=[job.id, "converted"]))
        self.assertEqual(response.status_code, 404)

    def test_download_invalid_kind_returns_404(self):
        self.client.force_login(self.user)
        job = ConvJob.objects.create(
            excel_file=SimpleUploadedFile("source.xlsx", make_source_workbook()),
        )
        response = self.client.get(reverse("download", args=[job.id, "bogus"]))
        self.assertEqual(response.status_code, 404)


def _nysc_row(cells):
    """Build a 37-wide source row with the given (0-indexed) columns set."""
    row = [""] * 37
    for index, value in cells.items():
        row[index] = value
    return row


def nysc_sample_rows():
    """Synthetic NY Supreme Court report rows: two judgments, the first with two
    co-defendants, the second a single company defendant."""
    return [
        _nysc_row({12: "Test County Clerk"}),                       # report header
        _nysc_row({2: "Index #"}),                                  # column header
        _nysc_row({2: 1001.0, 12: "2025/02/11 1032", 21: 5000.5,
                     30: "Plaintiff:", 36: "ACME BANK"}),
        _nysc_row({36: "100 MAIN ST\nNEW YORK, NY 10001"}),
        _nysc_row({30: "Defendant:", 36: "SMITH JOHN"}),
        _nysc_row({36: "200 OAK AVE\nYONKERS, NY 10701"}),
        _nysc_row({30: "Defendant:", 36: "SMITH JANE"}),
        _nysc_row({36: "200 OAK AVE\nYONKERS, NY 10701"}),
        _nysc_row({2: 1002.0, 12: "2025/02/12 0900", 21: 1234.0,
                     30: "Plaintiff:", 36: "BANK TWO"}),
        _nysc_row({36: "5 ELM ST\nRYE, NY 10580"}),
        _nysc_row({30: "Defendant:", 36: "WIDGETS LLC"}),
        _nysc_row({36: "9 PINE RD\nHARRISON, NY 10528"}),
    ]


class NyscParsingTests(TestCase):
    def test_extract_records_with_co_defendants(self):
        records = extract_nysc_records(nysc_sample_rows(), "Test County Clerk")
        self.assertEqual(len(records), 3)

        claim = [r for r in records if r["FilingNumber"] == "1001"]
        self.assertEqual([r["name"] for r in claim], ["SMITH JOHN", "SMITH JANE"])
        for record in claim:  # co-defendants share the claim's creditor + amount
            self.assertEqual(record["Creditor"], "ACME BANK")
            self.assertEqual(record["Judgment"], "$5,000.50")
            self.assertEqual(record["FilingDate"], "2025/02/11")
            self.assertEqual(record["FilingOffice"], "Test County Clerk")
        first = claim[0]
        self.assertEqual(
            (first["ADDRESS_1"], first["City"], first["State"], first["Zip"]),
            ("200 OAK AVE", "YONKERS", "NY", "10701"),
        )

        company = [r for r in records if r["FilingNumber"] == "1002"][0]
        self.assertEqual(company["name"], "WIDGETS LLC")  # company kept as-is
        self.assertEqual(company["Judgment"], "$1,234.00")

    def test_parse_address(self):
        self.assertEqual(
            _parse_nysc_address("889 JAMES ST\nPELHAM MAN, NY 10803"),
            ("889 JAMES ST", "PELHAM MAN", "NY", "10803"),
        )
        self.assertEqual(_parse_nysc_address(""), ("", "", "", ""))
        self.assertEqual(_parse_nysc_address("ONLY STREET"), ("ONLY STREET", "", "", ""))


@override_settings(MEDIA_ROOT=TEST_MEDIA_ROOT)
class NyscConversionTests(TestCase):
    @classmethod
    def tearDownClass(cls):
        super().tearDownClass()
        shutil.rmtree(TEST_MEDIA_ROOT, ignore_errors=True)

    def setUp(self):
        (TEST_MEDIA_ROOT / "excel_files").mkdir(parents=True, exist_ok=True)
        (TEST_MEDIA_ROOT / "conv_files").mkdir(parents=True, exist_ok=True)

    def test_convert_nysc_sheet_writes_mail_merge(self):
        job = ConvJob.objects.create(
            excel_file=SimpleUploadedFile("nysc.xls", b"dummy"), source_type="nysc"
        )
        with mock.patch.object(nysc, "_read_xls_rows", return_value=nysc_sample_rows()):
            self.assertTrue(nysc.convert_nysc_sheet(job))
        job.refresh_from_db()
        self.assertTrue(job.success)
        worksheet = load_workbook(TEST_MEDIA_ROOT / job.conv_file.name).active
        self.assertEqual(worksheet["A1"].value, "name")
        self.assertEqual(worksheet.max_row, 4)  # header + 3 defendants
        self.assertEqual(worksheet["A2"].value, "SMITH JOHN")
        self.assertEqual(worksheet["F2"].value, "ACME BANK")
        self.assertEqual(worksheet["G2"].value, "$5,000.50")


@override_settings(MEDIA_ROOT=TEST_MEDIA_ROOT)
class SourceRoutingTests(TestCase):
    @classmethod
    def tearDownClass(cls):
        super().tearDownClass()
        shutil.rmtree(TEST_MEDIA_ROOT, ignore_errors=True)

    def setUp(self):
        (TEST_MEDIA_ROOT / "excel_files").mkdir(parents=True, exist_ok=True)
        (TEST_MEDIA_ROOT / "conv_files").mkdir(parents=True, exist_ok=True)
        self.user = User.objects.create_user(username="router", password="password")

    def test_detect_source_lexisnexis(self):
        job = ConvJob.objects.create(
            excel_file=SimpleUploadedFile("ln.xlsx", make_source_workbook())
        )
        self.assertEqual(detect_source(job.excel_file.path), "lexisnexis")

    def test_detect_nysc_checks_extension_and_content(self):
        with mock.patch.object(nysc, "_read_xls_rows", return_value=nysc_sample_rows()):
            self.assertTrue(nysc.detect_nysc("/tmp/foo.xls"))
            self.assertFalse(nysc.detect_nysc("/tmp/foo.xlsx"))  # wrong extension

    def test_source_label(self):
        self.assertEqual(source_label("nysc"), "NY Supreme Court")
        self.assertEqual(source_label(""), "")

    def test_convert_job_routes_and_tags_lexisnexis(self):
        job = ConvJob.objects.create(
            excel_file=SimpleUploadedFile("ln.xlsx", make_source_workbook())
        )
        self.assertTrue(convert_job(job))
        job.refresh_from_db()
        self.assertTrue(job.success)
        self.assertEqual(job.source_type, "lexisnexis")  # detected + tagged

    def test_convert_job_unrecognized_format(self):
        job = ConvJob.objects.create(
            excel_file=SimpleUploadedFile("mystery.xlsx", b"not a workbook")
        )
        self.assertFalse(convert_job(job))
        job.refresh_from_db()
        self.assertFalse(job.success)
        self.assertIn("source format", (job.error or "").lower())

    def test_set_source_override(self):
        self.client.force_login(self.user)
        job = ConvJob.objects.create(
            excel_file=SimpleUploadedFile("ln.xlsx", make_source_workbook())
        )
        response = self.client.post(reverse("set_source", args=[job.id]), {"source_type": "nysc"})
        self.assertRedirects(response, reverse("jobs"))
        job.refresh_from_db()
        self.assertEqual(job.source_type, "nysc")
