"""Output schema and helpers shared by every source converter.

Each source format (LexisNexis, NY Supreme Court, ...) parses very different
input but writes the **same** converted mail-merge sheet, so the column schema
and the dollar formatter live here and are reused by all converters.
"""
import re

# Filing-detail columns (H-M of the output), in order.
FILING_FIELD_ORDER = (
    "FilingDate",
    "JudgmentType",
    "FilingNumber",
    "FilingDate2",
    "BookPage",
    "FilingOffice",
)

# Full converted-sheet header (columns A-M) and the matching column letters.
OUTPUT_HEADER = [
    "name", "ADDRESS_1", "City", "State", "Zip", "Creditor", "Judgment",
    *FILING_FIELD_ORDER,
]
OUTPUT_COLUMN_LETTERS = "ABCDEFGHIJKLM"


def format_dollars(value):
    """Format a dollar amount as ``$10,329.00``.

    Accepts a number (NY Supreme Court amounts arrive as floats) or a string
    that contains a number (e.g. ``$10,329`` or ``10329.5``). Returns ``""``
    when there is no parseable amount.
    """
    if value is None or value == "" or isinstance(value, bool):
        return ""
    if isinstance(value, (int, float)):
        try:
            return f"${float(value):,.2f}"
        except (ValueError, OverflowError):
            return ""
    match = re.search(r"\d[\d,]*(?:\.\d{1,2})?", str(value))
    if not match:
        return ""
    try:
        return f"${float(match.group(0).replace(',', '')):,.2f}"
    except ValueError:
        return ""


def write_mail_merge(records, out_path):
    """Write ``records`` (dicts keyed by OUTPUT_HEADER) to a new .xlsx at out_path."""
    from openpyxl.workbook import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    for col, label in enumerate(OUTPUT_HEADER, start=1):
        worksheet.cell(row=1, column=col, value=label)
    for row_index, record in enumerate(records, start=2):
        for col, key in enumerate(OUTPUT_HEADER, start=1):
            worksheet.cell(row=row_index, column=col, value=record.get(key, ""))
    workbook.save(out_path)
    workbook.close()
