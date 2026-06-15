from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from pathlib import Path
from django.core.files import File
from datetime import datetime
from django.utils import timezone
from django.conf import settings
import os
import re

from excel_conv.lib.mailmerge import (
    FILING_FIELD_ORDER,
    OUTPUT_COLUMN_LETTERS,
    OUTPUT_HEADER,
)


# --------------------------------------------------
def _format_judgment(filing_value):
    """Extract the judgment dollar amount from the source 'Filing' cell.

    The Filing column (D) holds multi-line text with a line like
    'Amount:$7,798'. Returns the value formatted as '$7,798.00', or '' when
    no amount is present.
    """
    if not filing_value:
        return ""
    match = re.search(r"Amount:\s*\$?([\d,]+(?:\.\d{1,2})?)", str(filing_value), re.IGNORECASE)
    if not match:
        return ""
    try:
        amount = float(match.group(1).replace(",", ""))
    except ValueError:
        return ""
    return f"${amount:,.2f}"


# --------------------------------------------------
def _parse_filing_fields(filing_value):
    """Parse the LexisNexis 'Filing' cell (column D) into named fields.

    The cell is multi-line -- labelled lines plus one unlabelled line for the
    judgment type, e.g.:
        Filing Date:1/12/2026
        Amount:$7,798
        SMALL CLAIMS JUDGMENT
        Filing Number:2025149620SP21
        Filing Date:1/12/2026
        Book/Page:35116/2236
        Filing Office:CIRCUIT COURT - CIVIL DIVISION - MIAMI, FL
    Returns a dict keyed by FILING_FIELD_ORDER. Any field that isn't present
    comes back as '' (an empty cell). 'Amount' is handled separately by the
    Judgment column.
    """
    fields = {key: "" for key in FILING_FIELD_ORDER}
    if not filing_value:
        return fields

    filing_dates = []
    for line in str(filing_value).splitlines():
        line = line.strip()
        if not line:
            continue
        if ":" in line:
            label, _, value = line.partition(":")
            label = label.strip().lower()
            value = value.strip()
            if label == "filing date":
                filing_dates.append(value)
            elif label == "filing number":
                fields["FilingNumber"] = value
            elif label == "book/page":
                fields["BookPage"] = value
            elif label == "filing office":
                fields["FilingOffice"] = value
            # 'Amount' and any other labelled lines are ignored here.
        elif not fields["JudgmentType"]:
            # The one unlabelled line is the judgment type (e.g. "CIVIL JUDGMENT").
            fields["JudgmentType"] = line

    if filing_dates:
        fields["FilingDate"] = filing_dates[0]
        if len(filing_dates) > 1:
            fields["FilingDate2"] = filing_dates[1]
    return fields


# --------------------------------------------------
def _select_data_worksheet(workbook):
    """Return the worksheet that holds the debtor records.

    LexisNexis exports keep the data on a 'Public Records Results List'
    sheet, but that sheet isn't always the active one -- the Florida
    exports open with an empty 'Sheet1' active, so a blind
    workbook.active read the wrong sheet and the conversion failed.
    Pick the first worksheet whose column A contains the 'No.' header;
    fall back to the active sheet.
    """
    for worksheet in workbook.worksheets:
        for row in range(1, min(worksheet.max_row, 50) + 1):
            if worksheet.cell(row=row, column=1).value == "No.":
                return worksheet
    return workbook.active


# --------------------------------------------------
def detect_lexisnexis(path):
    """True if path looks like a LexisNexis 'Public Records Results List' export."""
    if not str(path).lower().endswith(".xlsx"):
        return False
    try:
        workbook = load_workbook(path, data_only=True)
    except Exception:
        return False
    try:
        for worksheet in workbook.worksheets:
            if worksheet.title == "Public Records Results List":
                return True
            for row in range(1, min(worksheet.max_row, 50) + 1):
                if worksheet.cell(row=row, column=1).value == "No.":
                    return True
    finally:
        workbook.close()
    return False


# --------------------------------------------------
def convert_sheet(object):
    """convert_sheet function to test the conversion process"""
    
    # create the path to the temporary and converted Excel files
    out_file_path = object.excel_file.path.replace('excel_files', 'conv_files')
    out_file_path = str(out_file_path.replace('.xlsx', '_converted.xlsx'))
    temp_file_path = out_file_path.replace('.xlsx', '_temp.xlsx')

    # initialize the original workbook and worksheet
    original_workbook = Workbook()
    original_workbook = load_workbook(object.excel_file.path)
    original_worksheet = _select_data_worksheet(original_workbook)
    
    # initialize the converted workbook and worksheet
    converted_workbook = Workbook()
    converted_workbook.save(temp_file_path)
    converted_workbook = load_workbook(temp_file_path)
    converted_worksheet = converted_workbook.active
    
    # write the header row to the converted spreadsheet
    header_labels = OUTPUT_HEADER
    column_letters = OUTPUT_COLUMN_LETTERS
    for i, label in enumerate(header_labels):
        converted_worksheet[column_letters[i] + "1"] = label
    converted_workbook.save(temp_file_path)
    
    # we need to find the first and last debtor in the source file
    starting_row = None
    ending_row = None
    column_a = original_worksheet["A"]
    for item in column_a:
        if item.value == "No.":
            starting_row = item.row + 2
        if item.value == "Permissible Use:":
            ending_row = item.row - 3

    # if the markers are missing the file isn't in the expected layout;
    # fail gracefully instead of raising (which used to surface as a 500).
    if starting_row is None or ending_row is None:
        object.error = (
            "Could not find the debtor data range "
            "('No.' / 'Permissible Use:' markers) in the file."
        )
        object.success = False
        object.save()
        return False

    # now we're going to loop through the debtor rows and convert them
    i = starting_row
    new_file_row = 2
    # The first row of a claim carries the creditor + filing. Co-defendant
    # ("co-debtor") continuation rows that follow have only a name + address
    # and a blank creditor; carry the creditor/filing forward so each
    # co-debtor is emitted as its own row sharing the claim's creditor/amount.
    current_creditor = None
    current_filing = None

    while i <= ending_row:
        working_row = original_worksheet[i]
        new_file_working_row = converted_worksheet[new_file_row]

        data = {
            "debtor_name": working_row[1].value,
            "debtor_address": working_row[2].value,
            "filing": working_row[3].value,
            "creditor": working_row[4].value,
        }

        if data["creditor"]:
            current_creditor = data["creditor"]
            current_filing = data["filing"]
        else:
            data["creditor"] = current_creditor
            data["filing"] = current_filing

        # Process every row that has a debtor name once a creditor is in
        # scope -- this covers the primary debtor and each co-defendant.
        if data["debtor_name"] and data["creditor"]:
            try:
                # A creditor can be present while the name or address cell is
                # empty; skip those rows up front so calling .splitlines() on
                # None cannot crash the whole job (regression: job 516).
                if data["debtor_name"] is None or data["debtor_address"] is None:
                    print(f"Skipping row {i}: missing name or address.")
                    i += 1
                    continue

                # Process debtor name
                debtor_name_lines = data["debtor_name"].splitlines()
                if len(debtor_name_lines) >= 1 and debtor_name_lines[0]:
                    data["debtor_name"] = debtor_name_lines[0]
                else:
                    print(f"Skipping row {i}: debtor_name is empty or missing.")
                    i += 1
                    continue

                # Split debtor name into parts
                debtor_name_parts = data["debtor_name"].split(', ')
                if len(debtor_name_parts) == 2:
                    # "Last, First Middle" -> "First Middle Last"
                    last_name = debtor_name_parts[0]
                    first_name_parts = debtor_name_parts[1].split(' ')
                    first_name = first_name_parts[0]
                    middle_name = ' '.join(first_name_parts[1:]) if len(first_name_parts) > 1 else ''
                    # Join only the non-empty parts so a missing middle name
                    # doesn't leave a double space (e.g. "EARL  ROBLES").
                    data["full_name"] = " ".join(
                        part for part in (first_name, middle_name, last_name) if part
                    )
                else:
                    # No "Last, First" comma -> a company/organization (or a
                    # co-defendant business); keep the name as-is.
                    data["full_name"] = data["debtor_name"].strip()

                new_file_working_row[0].value = data["full_name"]

                # debtor_name_lines = data["debtor_name"].splitlines()
                # data["debtor_name"] = debtor_name_lines[0]
                #
                # last_name = data["debtor_name"].split(', ')[0]
                # first_name = data["debtor_name"].split(' ')[1]
                #
                # if len(data["debtor_name"].split(' ')) > 2 and data["creditor"] is not None:
                #     middle_name = data["debtor_name"].split(' ')[2]
                #     new_file_working_row[0].value = f"{first_name} {middle_name} {last_name}"
                # elif data["creditor"] is not None:
                #     middle_name = None
                #     new_file_working_row[0].value = f"{first_name} {last_name}"
                #
                # if middle_name:
                #     data["full_name"] = f"{first_name} {middle_name} {last_name}".strip()
                # else:
                #     data["full_name"] = f"{first_name} {last_name}".strip()

                # Process debtor address
                debtor_address_lines = data["debtor_address"].splitlines()

                if len(debtor_address_lines) == 3:
                    # Expected format with street address and city/state/zip
                    debtor_street_address = debtor_address_lines[0]
                    debtor_city_state_zip = debtor_address_lines[1]
                    debtor_county = debtor_address_lines[2]
                # elif len(debtor_address_lines) == 2:
                #     # Only city/state/zip provided
                #     debtor_street_address = ''
                #     debtor_city_state_zip = debtor_address_lines[0]
                else:
                    print(f"Skipping row {i}: debtor_address does not have expected number of lines.")
                    i += 1
                    continue

                # Split city, state, zip
                debtor_city_state_zip_parts = debtor_city_state_zip.split(', ')
                if len(debtor_city_state_zip_parts) == 2:
                    debtor_city = debtor_city_state_zip_parts[0]
                    state_zip_parts = debtor_city_state_zip_parts[1].split(' ')
                    if len(state_zip_parts) == 2:
                        debtor_state = state_zip_parts[0]
                        debtor_zip = state_zip_parts[1]
                    else:
                        print(f"Skipping row {i}: debtor_state and zip not in expected format.")
                        i += 1
                        continue
                else:
                    print(f"Skipping row {i}: debtor_city_state_zip not in expected format.")
                    i += 1
                    continue

                # debtor_street_address = data["debtor_address"].splitlines()[0]
                # debtor_city_state_zip = data["debtor_address"].splitlines()[1]
                # debtor_city = debtor_city_state_zip.split(', ')[0]
                # debtor_state = debtor_city_state_zip.split(', ')[1].split(' ')[0]
                # debtor_zip = debtor_city_state_zip.split(', ')[1].split(' ')[1]

                data.update(
                    {
                        "debtor_street_address": debtor_street_address,
                        "debtor_city": debtor_city,
                        "debtor_state": debtor_state,
                        "debtor_zip": debtor_zip,
                        "creditor": data["creditor"],
                    }
                )

                if data["creditor"] is not None:
                    new_file_working_row[1].value = debtor_street_address
                    new_file_working_row[2].value = debtor_city
                    new_file_working_row[3].value = debtor_state
                    new_file_working_row[4].value = debtor_zip
                    new_file_working_row[5].value = data["creditor"]
                    new_file_working_row[6].value = _format_judgment(data["filing"])
                    # Filing-detail columns (H-M); missing fields stay blank.
                    filing_fields = _parse_filing_fields(data["filing"])
                    for offset, key in enumerate(FILING_FIELD_ORDER, start=7):
                        new_file_working_row[offset].value = filing_fields[key]

                # for index, (key, value) in enumerate(data.items()):
                #     new_file_working_row[index].value = value

                new_file_row += 1
            except Exception as e:
                # Skip this row rather than aborting the whole job: one
                # malformed record must not block every other one.
                print(f"Skipping row {i}: error processing row: {e}")

        i += 1

    # Save once after all rows are processed. Saving inside the loop made this
    # O(n^2): a ~1500-row file took ~35s and timed out the gunicorn worker
    # (the "error/blank page" the client reported). One save is ~linear.
    converted_workbook.save(temp_file_path)

    # close the workbooks
    original_workbook.close()
    converted_workbook.close()
    
    # Create and write content to the new file
    with open(out_file_path, "wb") as file:
    # Write the content you want into the file (e.g., by copying from an existing file)
        with open(temp_file_path, "rb") as existing_file:
            file.write(existing_file.read())

    # Assign the path of the newly created file to the FileField
    # object.conv_file.name = object.excel_file.name.replace('.xlsx', '_converted.xlsx').replace('excel_files', 'conv_files') 
    object.conv_file.name = object.excel_file.name.replace('.xlsx', '_converted.xlsx')
    object.conv_file.name = object.conv_file.name.replace('excel_files', 'conv_files')
    
    # Assign a date
    object.conv_at = timezone.now()
    
    # Assign a success status
    object.success = True

    # Save the excel_file instance to persist the changes
    object.save()
    
    # delete the temp file
    os.remove(temp_file_path)
    
    # return status
    return object.success